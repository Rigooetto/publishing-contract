import csv
import io
import json as _json
import os
import datetime

from flask import Blueprint, render_template_string, request, redirect, url_for, flash, make_response, current_app

from extensions import db
from models import Work, WorkWriter
from utils import auth_required, role_required, FULL_ACCESS_ROLES
from ui import REGISTRATION_REPORT_HTML

_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "template")

bp = Blueprint("registration_report", __name__)


def _works_by_status(status):
    return (
        Work.query
        .filter_by(registration_status=status)
        .order_by(Work.title)
        .all()
    )


def _build_report():
    new_works       = _works_by_status("new")
    submitted_works = _works_by_status("submitted")
    confirmed_works = _works_by_status("confirmed")
    return new_works, submitted_works, confirmed_works


# ── Routes ────────────────────────────────────────────────────────────────────

@bp.route("/registration-report")
def registration_report():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    new_works, submitted_works, confirmed_works = _build_report()
    tab = request.args.get("tab", "new")

    stats = dict(
        new=len(new_works),
        submitted=len(submitted_works),
        confirmed=len(confirmed_works),
        total=len(new_works) + len(submitted_works) + len(confirmed_works),
    )

    return render_template_string(
        REGISTRATION_REPORT_HTML,
        new_works=new_works,
        submitted_works=submitted_works,
        confirmed_works=confirmed_works,
        stats=stats,
        tab=tab,
    )


@bp.route("/registration-report/mark-submitted", methods=["POST"])
def mark_submitted():
    """Mark selected (or all new) works as submitted."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    work_ids = request.form.getlist("work_ids", type=int)
    if not work_ids:
        flash("No works selected.", "error")
        return redirect(url_for("registration_report.registration_report", tab="new"))

    updated = (
        Work.query
        .filter(Work.id.in_(work_ids), Work.registration_status == "new")
        .all()
    )
    for w in updated:
        w.registration_status = "submitted"

    try:
        db.session.commit()
        flash(f"{len(updated)} work(s) marked as submitted.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {e}", "error")

    return redirect(url_for("registration_report.registration_report", tab="new"))


@bp.route("/registration-report/mark-new", methods=["POST"])
def mark_new():
    """Revert submitted works back to new (e.g. if a submission was recalled)."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    work_ids = request.form.getlist("work_ids", type=int)
    if not work_ids:
        flash("No works selected.", "error")
        return redirect(url_for("registration_report.registration_report", tab="submitted"))

    updated = (
        Work.query
        .filter(Work.id.in_(work_ids), Work.registration_status == "submitted")
        .all()
    )
    for w in updated:
        w.registration_status = "new"

    try:
        db.session.commit()
        flash(f"{len(updated)} work(s) moved back to New.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {e}", "error")

    return redirect(url_for("registration_report.registration_report", tab="submitted"))


@bp.route("/registration-report/export-csv")
def export_csv():
    """Export new or submitted works as CSV for sending to PRO/MLC."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    status = request.args.get("status", "new")
    if status not in ("new", "submitted"):
        status = "new"

    works = _works_by_status(status)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Work Title", "ISWC", "Contract Date",
        "Writer 1", "Writer 1 IPI", "Writer 1 PRO", "Writer 1 Split %",
        "Writer 2", "Writer 2 IPI", "Writer 2 PRO", "Writer 2 Split %",
        "Publisher 1", "Publisher 1 IPI",
        "MRI Song ID", "AKA Titles", "Registration Status",
    ])

    for w in works:
        writers_data = []
        publishers_data = []
        for ww in w.work_writers:
            writers_data.append((
                ww.writer.full_name,
                ww.writer.ipi or "",
                ww.writer.pro or "",
                ww.writer_percentage,
            ))
            if ww.publisher:
                publishers_data.append((ww.publisher, ww.publisher_ipi or ""))

        # Pad to at least 2 writer slots
        while len(writers_data) < 2:
            writers_data.append(("", "", "", ""))
        while len(publishers_data) < 1:
            publishers_data.append(("", ""))

        aka_str = "; ".join(a.title for a in w.aka_titles)

        row = [
            w.title,
            w.iswc or "",
            w.contract_date.strftime("%m/%d/%Y") if w.contract_date else "",
            writers_data[0][0], writers_data[0][1], writers_data[0][2], writers_data[0][3],
            writers_data[1][0], writers_data[1][1], writers_data[1][2], writers_data[1][3],
            publishers_data[0][0], publishers_data[0][1],
            w.mri_song_id or "",
            aka_str,
            w.registration_status,
        ]
        writer.writerow(row)

    content = output.getvalue()
    date_str = datetime.date.today().strftime("%Y%m%d")
    filename = f"registration_{status}_{date_str}.csv"

    resp = make_response(content)
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


# ── Smart Excel export helpers ────────────────────────────────────────────────

def _works_for_registration():
    """
    Returns (mlc_needs, mri_needs) — lists of Work objects that need registration.
      mlc_needs: in MRI but not MLC (mri_only) + in neither (unregistered)
      mri_needs: in MLC but not MRI (mlc_only) + in neither (unregistered)
    """
    from blueprints.mechanical_audit import _build_audit
    matched_both, mlc_only, mri_only, unregistered, orphaned, _mlc, _mri = _build_audit()

    seen_mlc = set()
    seen_mri = set()
    mlc_needs = []
    mri_needs = []

    for e in (mri_only + unregistered):
        w = e["work"]
        if w.id not in seen_mlc and w.registration_status == "new":
            seen_mlc.add(w.id)
            mlc_needs.append(w)

    for e in (mlc_only + unregistered):
        w = e["work"]
        if w.id not in seen_mri and w.registration_status == "new":
            seen_mri.add(w.id)
            mri_needs.append(w)

    return mlc_needs, mri_needs


def _track_info(work_id):
    """Return (rec_title, rec_artist, rec_isrc, rec_label, upc) for first linked track."""
    from models import Track, TrackWork, Release
    tracks = (Track.query
              .join(TrackWork, TrackWork.track_id == Track.id)
              .filter(TrackWork.work_id == work_id)
              .all())
    if not tracks:
        return "", "", "", "", ""
    t = tracks[0]
    try:
        rec_artist = ", ".join(a for a in _json.loads(t.artists or "[]") if a)
    except Exception:
        rec_artist = t.artists or ""
    upc = t.release.upc if t.release else ""
    return t.primary_title or "", rec_artist, t.isrc or "", t.track_label or "", upc or ""


# ── Smart Excel export routes ─────────────────────────────────────────────────

@bp.route("/registration-report/export-mlc")
def export_mlc():
    """MLC Bulk Work export — only works not yet registered at MLC."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        import openpyxl
        from openpyxl import load_workbook

        mlc_works, _ = _works_for_registration()

        template_path = os.path.join(_TEMPLATE_DIR, "MLCBulkWork_V1.2-2.xlsx")
        wb = load_workbook(template_path)
        ws = wb["Format"]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        row_idx = 2
        for work in mlc_works:
            wws = WorkWriter.query.filter_by(work_id=work.id).all()
            rec_title, rec_artist, rec_isrc, rec_label, _ = _track_info(work.id)

            first_writer = True
            for ww in wws:
                wr = ww.writer
                ws.cell(row=row_idx, column=1).value  = work.title if first_writer else None
                ws.cell(row=row_idx, column=2).value  = None  # MLC Song Code (assigned by MLC)
                ws.cell(row=row_idx, column=3).value  = f"LM{work.id:06d}"
                ws.cell(row=row_idx, column=4).value  = work.iswc or None
                ws.cell(row=row_idx, column=5).value  = work.aka_title or None
                ws.cell(row=row_idx, column=6).value  = work.aka_title_type_code or None
                ws.cell(row=row_idx, column=7).value  = wr.last_names or None
                ws.cell(row=row_idx, column=8).value  = wr.first_name or None
                ws.cell(row=row_idx, column=9).value  = wr.ipi or None
                ws.cell(row=row_idx, column=10).value = ww.writer_role_code or "CA"
                ws.cell(row=row_idx, column=11).value = None  # MLC Publisher Number
                ws.cell(row=row_idx, column=12).value = ww.publisher or None
                ws.cell(row=row_idx, column=13).value = ww.publisher_ipi or None
                ws.cell(row=row_idx, column=14).value = None
                ws.cell(row=row_idx, column=15).value = ww.administrator_name or None
                ws.cell(row=row_idx, column=16).value = ww.administrator_ipi or None
                ws.cell(row=row_idx, column=17).value = ww.writer_percentage or None
                ws.cell(row=row_idx, column=18).value = rec_title or None
                ws.cell(row=row_idx, column=19).value = rec_artist or None
                ws.cell(row=row_idx, column=20).value = rec_isrc or None
                ws.cell(row=row_idx, column=21).value = rec_label or None
                row_idx += 1
                first_writer = False

            if not wws:
                ws.cell(row=row_idx, column=1).value  = work.title
                ws.cell(row=row_idx, column=3).value  = f"LM{work.id:06d}"
                ws.cell(row=row_idx, column=4).value  = work.iswc or None
                ws.cell(row=row_idx, column=18).value = rec_title or None
                ws.cell(row=row_idx, column=19).value = rec_artist or None
                ws.cell(row=row_idx, column=20).value = rec_isrc or None
                row_idx += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"MLC_Registration_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        resp = make_response(output.read())
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    except Exception as e:
        current_app.logger.error("MLC smart export error: %s", e)
        flash("Error generating MLC export: " + str(e), "error")
        return redirect(url_for("registration_report.registration_report"))


@bp.route("/registration-report/export-mri")
def export_mri():
    """Music Reports catalog export — only works not yet registered at MRI."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        from blueprints.export_helpers import open_xls_template

        _, mri_works = _works_for_registration()

        template_path = os.path.join(_TEMPLATE_DIR, "MusicReportspublishing_catalog_template-3.xls")
        wb_xls = open_xls_template(template_path)
        ws = wb_xls.get_sheet(0)  # "Catalog Template" — row 0 is the header, preserved from template

        row_idx = 1
        for work in mri_works:
            wws = WorkWriter.query.filter_by(work_id=work.id).all()
            rec_title, rec_artist, rec_isrc, rec_label, upc = _track_info(work.id)

            first_writer = True
            for ww in wws:
                wr = ww.writer
                ws.write(row_idx, 0,  work.title if first_writer else "")
                ws.write(row_idx, 1,  work.aka_title or "")
                ws.write(row_idx, 2,  work.mri_song_id or "")
                ws.write(row_idx, 3,  f"LM{work.id:06d}")
                ws.write(row_idx, 4,  work.iswc or "")
                ws.write(row_idx, 5,  wr.last_names or "")
                ws.write(row_idx, 6,  wr.first_name or "")
                ws.write(row_idx, 7,  wr.middle_name or "")
                ws.write(row_idx, 8,  wr.pro or "")
                ws.write(row_idx, 9,  wr.ipi or "")
                ws.write(row_idx, 10, "Y")
                ws.write(row_idx, 11, ww.writer_percentage or 0)
                ws.write(row_idx, 12, ww.writer_role_code or "CA")
                ws.write(row_idx, 13, ww.publisher or "")
                ws.write(row_idx, 14, "")  # Publisher PRO (not stored per-publisher)
                ws.write(row_idx, 15, ww.publisher_ipi or "")
                ws.write(row_idx, 16, "Y")
                ws.write(row_idx, 17, ww.administrator_name or "")
                ws.write(row_idx, 18, ww.writer_percentage or 0)
                ws.write(row_idx, 19, ww.territory_controlled or "World")
                ws.write(row_idx, 20, "")  # Territory exclusions
                ws.write(row_idx, 21, "")  # Publisher address
                ws.write(row_idx, 22, "")  # Publisher contact
                ws.write(row_idx, 23, rec_artist if first_writer else "")
                ws.write(row_idx, 24, rec_label  if first_writer else "")
                ws.write(row_idx, 25, rec_isrc   if first_writer else "")
                ws.write(row_idx, 26, upc        if first_writer else "")
                row_idx += 1
                first_writer = False

            if not wws:
                ws.write(row_idx, 0, work.title)
                ws.write(row_idx, 2, work.mri_song_id or "")
                ws.write(row_idx, 3, f"LM{work.id:06d}")
                ws.write(row_idx, 4, work.iswc or "")
                row_idx += 1

        output = io.BytesIO()
        wb_xls.save(output)
        output.seek(0)
        filename = f"MRI_Registration_{datetime.date.today().strftime('%Y%m%d')}.xls"
        resp = make_response(output.read())
        resp.headers["Content-Type"] = "application/vnd.ms-excel"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    except Exception as e:
        current_app.logger.error("MRI smart export error: %s", e)
        flash("Error generating Music Reports export: " + str(e), "error")
        return redirect(url_for("registration_report.registration_report"))


@bp.route("/registration-report/export-soundexchange")
def export_soundexchange():
    """SoundExchange ISRC Ingest export — only tracks not yet registered."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        from openpyxl import load_workbook
        from blueprints.neighboring_rights_audit import _build_audit as _nr_build_audit

        _matched, unregistered, _orphaned, _total = _nr_build_audit()

        template_path = os.path.join(_TEMPLATE_DIR, "Sound Exchange ISRC Ingest Form.xlsx")
        wb = load_workbook(template_path)
        ws = wb["Form"]

        data_start = 11
        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        from models import TrackWork, Work as _Work
        row_idx = data_start
        for entry in unregistered:
            t   = entry["track"]
            rel = entry["release"]

            # Skip if all linked works are already submitted or confirmed
            linked_works = (_Work.query
                            .join(TrackWork, TrackWork.work_id == _Work.id)
                            .filter(TrackWork.track_id == t.id)
                            .all())
            if linked_works and all(w.registration_status != "new" for w in linked_works):
                continue

            try:
                artist = ", ".join(a for a in _json.loads(t.artists or "[]") if a)
            except Exception:
                artist = t.artists or ""
            if not artist:
                try:
                    artist = ", ".join(a for a in _json.loads(rel.artists or "[]") if a)
                except Exception:
                    artist = rel.artists or ""

            ws.cell(row=row_idx, column=1).value  = artist
            ws.cell(row=row_idx, column=2).value  = t.primary_title
            ws.cell(row=row_idx, column=3).value  = t.isrc or ""
            ws.cell(row=row_idx, column=4).value  = "Copyright Owner"
            ws.cell(row=row_idx, column=5).value  = 100
            ws.cell(row=row_idx, column=6).value  = (rel.release_date.strftime("%m/%d/%Y")
                                                      if rel.release_date else "")
            ws.cell(row=row_idx, column=7).value  = ""
            ws.cell(row=row_idx, column=8).value  = ""
            ws.cell(row=row_idx, column=9).value  = ""
            ws.cell(row=row_idx, column=10).value = t.duration or ""
            ws.cell(row=row_idx, column=11).value = t.genre or ""
            ws.cell(row=row_idx, column=12).value = (t.recording_date.strftime("%m/%d/%Y")
                                                      if t.recording_date else "")
            ws.cell(row=row_idx, column=13).value = t.country_of_recording or "US"
            ws.cell(row=row_idx, column=14).value = ""
            ws.cell(row=row_idx, column=15).value = "US"
            row_idx += 1

        from blueprints.export_helpers import stitch_xlsx_assets
        opx_buf = io.BytesIO()
        wb.save(opx_buf)
        fixed_bytes = stitch_xlsx_assets(template_path, opx_buf.getvalue())
        filename = f"SoundExchange_Registration_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        resp = make_response(fixed_bytes)
        resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return resp
    except Exception as e:
        current_app.logger.error("SoundExchange smart export error: %s", e)
        flash("Error generating SoundExchange export: " + str(e), "error")
        return redirect(url_for("registration_report.registration_report"))
