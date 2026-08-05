import csv
import datetime
import glob
import io
import json
import os
import re

import json as _json

from flask import Blueprint, render_template_string, request, redirect, url_for, flash, send_file
from flask import current_app
from sqlalchemy import func as _func

from extensions import db
from models import Work, WorkWriter, ProRegistration, WorkAKA, PublisherConfig, Track, TrackWork
from utils import auth_required, paginate_list, role_required, FULL_ACCESS_ROLES, normalize_for_match
from ui import MECHANICAL_AUDIT_HTML

bp = Blueprint("mechanical_audit", __name__)

# ── Paths ─────────────────────────────────────────────────────────────────────

_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "template")
_UPLOAD_DIR   = os.path.join(os.path.dirname(__file__), "..", "uploads", "mechanical_catalogs")

# Required columns (uppercased) to validate uploads
_REQUIRED_COLS = {
    "mlc": {"PRIMARY TITLE", "MLC SONG CODE"},
    "mri": {"SONG TITLE", "MRI ID"},
}

_ACCEPTED_EXTS = {".csv", ".xlsx", ".xls"}


def _find_template(pattern):
    """Find the newest file in template dir matching a glob pattern."""
    matches = sorted(glob.glob(os.path.join(_TEMPLATE_DIR, pattern)))
    return matches[-1] if matches else ""


# ── File I/O ──────────────────────────────────────────────────────────────────

def _norm_key(k):
    """Strip trailing * and whitespace, uppercase — for consistent header lookup."""
    return re.sub(r"\s*\*+\s*$", "", str(k or "")).strip().upper()


def _rows_from_xlsx(content_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes))
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return []
    headers = [_norm_key(h) for h in raw[0]]
    return [dict(zip(headers, row)) for row in raw[1:]]


def _rows_from_xls(content_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=content_bytes)
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return []
    headers = [_norm_key(ws.cell_value(0, c)) for c in range(ws.ncols)]
    return [dict(zip(headers, [ws.cell_value(r, c) for c in range(ws.ncols)]))
            for r in range(1, ws.nrows)]


def _rows_from_csv(content_str):
    reader = csv.DictReader(io.StringIO(content_str))
    return [{_norm_key(k): v for k, v in row.items()} for row in reader]


def _read_file(content_bytes, filename):
    ext = os.path.splitext(filename.lower())[1]
    if ext == ".xlsx":
        return _rows_from_xlsx(content_bytes)
    if ext == ".xls":
        return _rows_from_xls(content_bytes)
    return _rows_from_csv(content_bytes.decode("utf-8-sig", errors="replace"))


# ── Upload metadata ───────────────────────────────────────────────────────────

def _uploaded_path(source):
    meta = _read_meta(source)
    if meta and meta.get("ext"):
        p = os.path.join(_UPLOAD_DIR, f"{source}_catalog{meta['ext']}")
        if os.path.exists(p):
            return p
    for ext in _ACCEPTED_EXTS:
        p = os.path.join(_UPLOAD_DIR, f"{source}_catalog{ext}")
        if os.path.exists(p):
            return p
    return ""


def _active_path(source):
    up = _uploaded_path(source)
    if up:
        return up
    # Fall back to bundled template files
    patterns = {
        "mlc": "mlc_work_report*.csv",
        "mri": "Music Reports*.csv",
    }
    return _find_template(patterns.get(source, ""))


def _meta_path(source):
    return os.path.join(_UPLOAD_DIR, f"{source}_meta.json")


def _read_meta(source):
    path = _meta_path(source)
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return None


def _write_meta(source, original_name, ext, work_count):
    os.makedirs(_UPLOAD_DIR, exist_ok=True)
    with open(_meta_path(source), "w") as f:
        json.dump({
            "original_name": original_name,
            "ext": ext,
            "uploaded_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "work_count": work_count,
        }, f)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(title):
    return normalize_for_match(title)


def _s(val):
    return str(val).strip() if val is not None else ""


def _parse_date(raw, *fmts):
    for fmt in fmts:
        try:
            return datetime.datetime.strptime(raw.strip(), fmt).date()
        except (ValueError, AttributeError):
            pass
    return datetime.date.today()


# ── CSV parsers ───────────────────────────────────────────────────────────────

def _parse_mlc():
    """
    Actual MLC work report export columns (after uppercasing):
    MLC SONG CODE, PRIMARY TITLE, ISWC, MEMBER SONG IDS, ARTISTS,
    PARTY ID, PARTY NAME, PARTY IPI, PARTY ROLE, COLLECTIBLE SHARE, ...
    Multiple rows per work (one per party).
    """
    path = _active_path("mlc")
    if not path or not os.path.exists(path):
        return {}

    with open(path, "rb") as f:
        rows = _read_file(f.read(), path)

    by_code = {}   # keyed by MLC Song Code (most reliable)
    for row in rows:
        title    = _s(row.get("PRIMARY TITLE", ""))
        mlc_code = _s(row.get("MLC SONG CODE", ""))
        iswc     = _s(row.get("ISWC", ""))
        artists  = _s(row.get("ARTISTS", ""))
        party    = _s(row.get("PARTY NAME", ""))
        role     = _s(row.get("PARTY ROLE", "")).lower()
        if not title:
            continue

        key = mlc_code if mlc_code else _norm(title)
        if key not in by_code:
            by_code[key] = dict(title=title, mlc_song_code=mlc_code,
                                iswc=iswc, writers=[], publisher="", artist=artists)
        w = by_code[key]
        if iswc and not w["iswc"]:
            w["iswc"] = iswc
        if artists and not w["artist"]:
            w["artist"] = artists

        is_writer    = "composer" in role or ("writer" in role and "publisher" not in role)
        is_publisher = "publisher" in role or "administrator" in role

        if is_writer and party and party not in w["writers"]:
            w["writers"].append(party)
        elif is_publisher and not w["publisher"]:
            w["publisher"] = party

    # Re-index by normalized title for DB matching
    return {_norm(v["title"]): v for v in by_code.values()}


def _parse_mri():
    """
    Actual Music Reports / Songdex portal export columns (after uppercasing):
    MRI ID, SONG TITLE, PUBLISHER(S), COMPOSER(S), US SHARE, REVISED SHARE,
    NOT OUR SONG, LAST UPDATED
    One row per work.
    """
    path = _active_path("mri")
    if not path or not os.path.exists(path):
        return {}

    with open(path, "rb") as f:
        rows = _read_file(f.read(), path)

    works = {}
    for row in rows:
        title     = _s(row.get("SONG TITLE", ""))
        mri_id    = _s(row.get("MRI ID", ""))
        publisher = _s(row.get("PUBLISHER(S)", ""))
        composer  = _s(row.get("COMPOSER(S)", ""))
        if not title:
            continue
        key = _norm(title)
        if key not in works:
            works[key] = dict(title=title, mri_song_id=mri_id, iswc="",
                              writers=[], publisher=publisher, artist="")
        w = works[key]
        if mri_id and not w["mri_song_id"]:
            w["mri_song_id"] = mri_id
        if composer and composer not in w["writers"]:
            w["writers"].append(composer)
    return works


# ── Audit builder ─────────────────────────────────────────────────────────────

def _build_audit():
    mlc = _parse_mlc()
    mri = _parse_mri()

    all_works = Work.query.order_by(Work.title).all()
    db_keys   = {_norm(w.title): w for w in all_works}
    # Secondary lookup by plain lowercase — catches edge cases _norm misses
    db_keys_lower = {w.title.lower(): w for w in all_works}

    # Include AKA keys — re-normalize at runtime so stale uppercase stored values match
    for w in all_works:
        for aka in w.aka_titles:
            nk = _norm(aka.title)
            if nk not in db_keys:
                db_keys[nk] = w
            lk = aka.title.lower()
            if lk not in db_keys_lower:
                db_keys_lower[lk] = w

    matched_both = []
    mlc_only     = []
    mri_only     = []
    unregistered = []

    for w in all_works:
        key = _norm(w.title)
        aka_keys = [_norm(aka.title) for aka in w.aka_titles]

        m = mlc.get(key) or next((mlc.get(k) for k in aka_keys if mlc.get(k)), None)
        r = mri.get(key) or next((mri.get(k) for k in aka_keys if mri.get(k)), None)

        iswcs = set(filter(None, [(m or {}).get("iswc"), (r or {}).get("iswc")]))
        iswc_conflict  = len(iswcs) > 1
        suggested_iswc = next(iter(iswcs), "") if not w.iswc else ""

        # Suggest a source title only when it is mixed-case AND has diacritics the
        # DB title is missing — never when the source is just ALL CAPS or same title.
        def _worth_suggesting(src_title, db_title):
            if not src_title or src_title == db_title:
                return False
            if src_title.lower() == db_title.lower():
                return False
            if normalize_for_match(src_title) == src_title.lower():
                return False
            return True

        src_titles = [
            src["title"] for src in [m, r]
            if src and _worth_suggesting(src.get("title", ""), w.title)
        ]
        suggested_title = src_titles[0] if src_titles else ""

        entry = dict(work=w, mlc=m, mri=r,
                     suggested_iswc=suggested_iswc,
                     iswc_conflict=iswc_conflict,
                     iswcs_found=iswcs,
                     suggested_title=suggested_title)
        if m and r:
            matched_both.append(entry)
        elif m:
            mlc_only.append(entry)
        elif r:
            mri_only.append(entry)
        else:
            unregistered.append(entry)

    orphaned = []
    for source, src_dict in [("MLC", mlc), ("MRI", mri)]:
        for key, v in src_dict.items():
            if key not in db_keys and v["title"].lower() not in db_keys_lower:
                orphaned.append(dict(source=source, **v))

    return matched_both, mlc_only, mri_only, unregistered, orphaned, mlc, mri


# ── Routes ────────────────────────────────────────────────────────────────────

@bp.route("/mechanical-audit")
def mechanical_audit():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    matched_both, mlc_only, mri_only, unregistered, orphaned, mlc, mri = _build_audit()

    mlc_unregistered = mri_only + unregistered   # not in MLC → needs MLC registration
    mri_unregistered = mlc_only + unregistered   # not in MRI → needs MRI registration

    stats = dict(
        mlc_total=len(mlc),
        mri_total=len(mri),
        db_total=len(matched_both) + len(mlc_only) + len(mri_only) + len(unregistered),
        matched_both=len(matched_both),
        mlc_only=len(mlc_only),
        mri_only=len(mri_only),
        unregistered=len(unregistered),
        orphaned=len(orphaned),
        mlc_unregistered=len(mlc_unregistered),
        mri_unregistered=len(mri_unregistered),
    )

    upload_meta = {"mlc": _read_meta("mlc"), "mri": _read_meta("mri")}
    tab  = request.args.get("tab", "matched_both")
    page = request.args.get("page", 1, type=int)

    if tab == "matched_both":
        pagination = paginate_list(matched_both, page)
        matched_both = pagination.items
    elif tab == "mlc_only":
        pagination = paginate_list(mlc_only, page)
        mlc_only = pagination.items
    elif tab == "mri_only":
        pagination = paginate_list(mri_only, page)
        mri_only = pagination.items
    elif tab == "unregistered":
        pagination = paginate_list(unregistered, page)
        unregistered = pagination.items
    elif tab == "mlc_unregistered":
        pagination = paginate_list(mlc_unregistered, page)
        mlc_unregistered = pagination.items
    elif tab == "mri_unregistered":
        pagination = paginate_list(mri_unregistered, page)
        mri_unregistered = pagination.items
    else:
        pagination = paginate_list(orphaned, page)
        orphaned = pagination.items

    return render_template_string(
        MECHANICAL_AUDIT_HTML,
        matched_both=matched_both, mlc_only=mlc_only, mri_only=mri_only,
        unregistered=unregistered, orphaned=orphaned,
        mlc_unregistered=mlc_unregistered, mri_unregistered=mri_unregistered,
        stats=stats, tab=tab, upload_meta=upload_meta, pagination=pagination,
    )


@bp.route("/mechanical-audit/upload", methods=["POST"])
def upload_catalog():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    source = request.form.get("source", "").lower().strip()
    file   = request.files.get("file")

    if source not in ("mlc", "mri"):
        flash("Invalid source.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))
    if not file or not file.filename:
        flash("No file selected.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    ext = os.path.splitext(file.filename.lower())[1]
    if ext not in _ACCEPTED_EXTS:
        flash("Only .csv, .xlsx, or .xls files are accepted.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    content_bytes = file.read()
    try:
        rows = _read_file(content_bytes, file.filename)
    except Exception as e:
        flash(f"Could not read file: {e}", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    if not rows:
        flash("The uploaded file appears to be empty.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    actual_cols = set(rows[0].keys())
    required    = _REQUIRED_COLS[source]
    if not required.issubset(actual_cols):
        missing = required - actual_cols
        flash(f"File missing expected columns: {', '.join(missing)}. "
              f"Make sure this is a {source.upper()} catalog export.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    title_col  = "PRIMARY TITLE" if source == "mlc" else "SONG TITLE"
    work_count = len({_s(r.get(title_col, "")) for r in rows
                      if _s(r.get(title_col, ""))})

    # Remove old files for this source
    for old_ext in _ACCEPTED_EXTS:
        old = os.path.join(_UPLOAD_DIR, f"{source}_catalog{old_ext}")
        if os.path.exists(old):
            os.remove(old)

    os.makedirs(_UPLOAD_DIR, exist_ok=True)
    with open(os.path.join(_UPLOAD_DIR, f"{source}_catalog{ext}"), "wb") as f_out:
        f_out.write(content_bytes)

    _write_meta(source, file.filename, ext, work_count)
    flash(f"{source.upper()} catalog updated — {work_count} works loaded.", "success")
    return redirect(url_for("mechanical_audit.mechanical_audit"))


@bp.route("/mechanical-audit/apply", methods=["POST"])
def apply_sync():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    matched_both, mlc_only, mri_only, _u, _o, _m, _r = _build_audit()
    all_matched = matched_both + mlc_only + mri_only

    iswc_updated = mri_updated = mlc_created = skipped_iswc = confirmed_count = 0

    for entry in all_matched:
        w = entry["work"]
        m = entry["mlc"]
        r = entry["mri"]

        # Auto-confirm registration status
        if w.registration_status != "confirmed":
            w.registration_status = "confirmed"
            confirmed_count += 1

        if not w.iswc:
            if entry["iswc_conflict"]:
                skipped_iswc += 1
            elif entry["suggested_iswc"]:
                w.iswc = entry["suggested_iswc"]
                iswc_updated += 1

        if r and r.get("mri_song_id") and not w.mri_song_id:
            w.mri_song_id = r["mri_song_id"]
            mri_updated += 1

        if m and m.get("mlc_song_code"):
            exists = ProRegistration.query.filter_by(work_id=w.id, pro="MLC").first()
            if exists:
                if not exists.mlc_song_code:
                    exists.mlc_song_code = m["mlc_song_code"]
                    mlc_created += 1
            else:
                db.session.add(ProRegistration(
                    work_id=w.id, pro="MLC",
                    mlc_song_code=m["mlc_song_code"],
                    registered_at=datetime.date.today(),
                    registered_by="MLC CSV Import",
                    notes="Imported from MLC catalog export",
                ))
                mlc_created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {e}", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit"))

    parts = []
    if iswc_updated:    parts.append(f"{iswc_updated} ISWC numbers written")
    if mri_updated:     parts.append(f"{mri_updated} MRI Song IDs written")
    if mlc_created:     parts.append(f"{mlc_created} MLC Song Codes saved")
    if confirmed_count: parts.append(f"{confirmed_count} works marked confirmed")
    if skipped_iswc:    parts.append(f"{skipped_iswc} works skipped (ISWC conflict)")
    flash(", ".join(parts) + "." if parts else "Nothing to update.", "success")
    return redirect(url_for("mechanical_audit.mechanical_audit", tab="matched_both"))


@bp.route("/mechanical-audit/apply-title", methods=["POST"])
def apply_title():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    work_id   = request.form.get("work_id", type=int)
    new_title = (request.form.get("new_title") or "").strip()
    if not work_id or not new_title:
        flash("Invalid request.", "error")
        return redirect(url_for("mechanical_audit.mechanical_audit", tab="matched_both"))

    work = Work.query.get_or_404(work_id)
    old_title = work.title
    work.title = new_title
    work.normalized_title = _norm(new_title)
    try:
        db.session.commit()
        flash(f'Title updated: "{old_title}" → "{new_title}".', "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {e}", "error")
    return redirect(url_for("mechanical_audit.mechanical_audit", tab="matched_both"))


@bp.route("/mechanical-audit/export/mlc-unregistered")
def export_mlc_unregistered():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        import openpyxl

        _mb, mlc_only, mri_only, unregistered, _o, _mlc, _mri = _build_audit()
        entries = mri_only + unregistered

        template_path = os.path.join(_TEMPLATE_DIR, "MLCBulkWork_V1.2-2.xlsx")
        wb = openpyxl.load_workbook(template_path)
        ws = wb["Format"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        row_idx = 2
        for entry in entries:
            work = entry["work"]
            writers = WorkWriter.query.filter_by(work_id=work.id).all()
            if not writers:
                ws.cell(row=row_idx, column=1).value = work.title
                ws.cell(row=row_idx, column=3).value = f"LM{work.id:06d}"
                ws.cell(row=row_idx, column=4).value = work.iswc or None
                row_idx += 1
                continue
            first = True
            for ww in writers:
                w = ww.writer
                pub_config = PublisherConfig.query.filter(
                    _func.lower(PublisherConfig.publisher_name) == (ww.publisher or "").lower()
                ).first()
                mlc_pub_num = pub_config.mlc_publisher_number if pub_config else ""

                tracks = (Track.query
                          .join(TrackWork, TrackWork.track_id == Track.id)
                          .filter(TrackWork.work_id == work.id).all())
                rec_title  = tracks[0].primary_title if tracks else ""
                rec_isrc   = tracks[0].isrc if tracks else ""
                rec_label  = tracks[0].track_label if tracks else ""
                rec_artist = ""
                if tracks:
                    try:
                        rec_artist = ", ".join(_json.loads(tracks[0].artists or "[]"))
                    except Exception:
                        pass

                ws.cell(row=row_idx, column=1).value  = work.title if first else None
                ws.cell(row=row_idx, column=2).value  = None
                ws.cell(row=row_idx, column=3).value  = f"LM{work.id:06d}"
                ws.cell(row=row_idx, column=4).value  = work.iswc or None
                ws.cell(row=row_idx, column=7).value  = w.last_names
                ws.cell(row=row_idx, column=8).value  = w.first_name
                ws.cell(row=row_idx, column=9).value  = w.ipi or None
                ws.cell(row=row_idx, column=10).value = ww.writer_role_code or "CA"
                ws.cell(row=row_idx, column=11).value = mlc_pub_num or None
                ws.cell(row=row_idx, column=12).value = ww.publisher or None
                ws.cell(row=row_idx, column=13).value = ww.publisher_ipi or None
                ws.cell(row=row_idx, column=15).value = ww.administrator_name or None
                ws.cell(row=row_idx, column=16).value = ww.administrator_ipi or None
                ws.cell(row=row_idx, column=17).value = ww.writer_percentage or None
                ws.cell(row=row_idx, column=18).value = rec_title or None
                ws.cell(row=row_idx, column=19).value = rec_artist or None
                ws.cell(row=row_idx, column=20).value = rec_isrc or None
                ws.cell(row=row_idx, column=21).value = rec_label or None
                row_idx += 1
                first = False

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"MLC_Unregistered_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(output, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True)
    except Exception as e:
        current_app.logger.error("MLC unregistered export error: %s", e)
        flash("Error generating export: " + str(e), "error")
        return redirect(url_for("mechanical_audit.mechanical_audit", tab="mlc_unregistered"))


@bp.route("/mechanical-audit/export/mri-unregistered")
def export_mri_unregistered():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        from blueprints.export_helpers import open_xls_template

        _mb, mlc_only, mri_only, unregistered, _o, _mlc, _mri = _build_audit()
        entries = mlc_only + unregistered  # not in MRI

        template_path = os.path.join(_TEMPLATE_DIR, "MusicReportspublishing_catalog_template-3.xls")
        wb = open_xls_template(template_path)
        ws = wb.get_sheet(0)

        _afinarte = ["songs of afinarte", "melodies of afinarte", "music of afinarte"]
        def _ctrl(pub):
            return any(a in (pub or "").lower() for a in _afinarte)

        row_idx = 1
        for entry in entries:
            work = entry["work"]
            writers = WorkWriter.query.filter_by(work_id=work.id).all()
            tracks = (Track.query
                      .join(TrackWork, TrackWork.track_id == Track.id)
                      .filter(TrackWork.work_id == work.id)
                      .all())
            rec_artist = rec_isrc = rec_label = upc = ""
            if tracks:
                try:
                    al = _json.loads(tracks[0].artists or "[]")
                    rec_artist = ", ".join(al)
                except Exception:
                    pass
                rec_isrc = tracks[0].isrc or ""
                rec_label = tracks[0].track_label or ""
                if tracks[0].release:
                    upc = tracks[0].release.upc or ""

            if not writers:
                ws.write(row_idx, 0, work.title)
                ws.write(row_idx, 3, f"LM{work.id:06d}")
                ws.write(row_idx, 4, work.iswc or "")
                row_idx += 1
                continue

            first = True
            for ww in writers:
                w = ww.writer
                pub_config = PublisherConfig.query.filter(
                    _func.lower(PublisherConfig.publisher_name) == (ww.publisher or "").lower()
                ).first()
                pub_address = pub_contact = pub_pro = ""
                if pub_config:
                    parts = [pub_config.address, pub_config.city]
                    if pub_config.state:
                        parts.append(pub_config.state)
                    if pub_config.zip_code:
                        parts.append(pub_config.zip_code)
                    pub_address = ", ".join(p for p in parts if p)
                    pub_contact = pub_config.contact_email or pub_config.contact_phone or ""
                    pub_pro = pub_config.pro or ""

                controlled = "Y" if _ctrl(ww.publisher) else "N"

                ws.write(row_idx, 0, work.title if first else "")
                ws.write(row_idx, 1, work.aka_title or "")
                ws.write(row_idx, 2, work.mri_song_id or "")
                ws.write(row_idx, 3, f"LM{work.id:06d}")
                ws.write(row_idx, 4, work.iswc or "")
                ws.write(row_idx, 5, w.last_names or "")
                ws.write(row_idx, 6, w.first_name or "")
                ws.write(row_idx, 7, w.middle_name or "")
                ws.write(row_idx, 8, w.pro or "")
                ws.write(row_idx, 9, w.ipi or "")
                ws.write(row_idx, 10, controlled)
                ws.write(row_idx, 11, ww.writer_percentage or 0)
                ws.write(row_idx, 12, ww.writer_role_code or "CA")
                ws.write(row_idx, 13, ww.publisher or "")
                ws.write(row_idx, 14, pub_pro)
                ws.write(row_idx, 15, ww.publisher_ipi or "")
                ws.write(row_idx, 16, controlled)
                ws.write(row_idx, 17, ww.administrator_name or "")
                ws.write(row_idx, 18, ww.writer_percentage or 0)
                ws.write(row_idx, 19, ww.territory_controlled or "World")
                ws.write(row_idx, 20, "")
                ws.write(row_idx, 21, pub_address)
                ws.write(row_idx, 22, pub_contact)
                ws.write(row_idx, 23, rec_artist if first else "")
                ws.write(row_idx, 24, rec_label if first else "")
                ws.write(row_idx, 25, rec_isrc if first else "")
                ws.write(row_idx, 26, upc if first else "")

                row_idx += 1
                first = False

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"MRI_Unregistered_{datetime.date.today().strftime('%Y%m%d')}.xls"
        return send_file(output, download_name=filename,
                         mimetype="application/vnd.ms-excel",
                         as_attachment=True)
    except Exception as e:
        current_app.logger.error("MRI unregistered export error: %s", e)
        flash("Error generating export: " + str(e), "error")
        return redirect(url_for("mechanical_audit.mechanical_audit", tab="mri_unregistered"))
