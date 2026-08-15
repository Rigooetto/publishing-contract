import io
import os
import datetime
import json as _json

_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "..", "template")

from flask import Blueprint, request, redirect, url_for, flash, render_template_string, send_file
from flask import current_app
from sqlalchemy import func as _func

from extensions import db
from models import (Work, WorkWriter, ProRegistration, PublisherConfig,
                    Release, Track, TrackWork)
from utils import auth_required, role_required, FULL_ACCESS_ROLES
from ui import REPORTS_INDEX_HTML, PUBLISHER_CONFIG_HTML, PRO_REGISTRATION_HTML, WORKS_REGISTRATION_HTML

bp = Blueprint("reports", __name__)

AFINARTE_PUBLISHERS = ["Songs of Afinarte", "Melodies of Afinarte", "Music of Afinarte"]


def _attach_track_info(work):
    tracks = (Track.query
              .join(TrackWork, TrackWork.track_id == Track.id)
              .filter(TrackWork.work_id == work.id)
              .all())
    work._tracks = tracks
    work._first_track = tracks[0] if tracks else None
    work._first_release = tracks[0].release if tracks and tracks[0].release else None
    try:
        work._artist_display = ", ".join(_json.loads(tracks[0].artists or "[]")) if tracks else ""
    except Exception:
        work._artist_display = ""


def _attach_track_info_bulk(works):
    """Single query for all works on the page instead of one query per work."""
    if not works:
        return
    work_ids = [w.id for w in works]
    rows = (db.session.query(TrackWork.work_id, Track)
            .join(Track, Track.id == TrackWork.track_id)
            .filter(TrackWork.work_id.in_(work_ids))
            .all())
    from collections import defaultdict
    tracks_by_work = defaultdict(list)
    for work_id, track in rows:
        tracks_by_work[work_id].append(track)
    for w in works:
        tracks = tracks_by_work.get(w.id, [])
        w._tracks = tracks
        w._first_track = tracks[0] if tracks else None
        w._first_release = tracks[0].release if tracks and tracks[0].release else None
        try:
            w._artist_display = ", ".join(_json.loads(tracks[0].artists or "[]")) if tracks else ""
        except Exception:
            w._artist_display = ""


def _is_controlled(publisher_name):
    if not publisher_name:
        return False
    return any(ap.lower() in publisher_name.lower() for ap in AFINARTE_PUBLISHERS)


# ── Publisher Config ──────────────────────────────────────────────────────────

@bp.route("/publisher-config", methods=["GET", "POST"])
def publisher_config():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    if request.method == "POST":
        names = request.form.getlist("publisher_name[]")
        pros = request.form.getlist("pro[]")
        ipis = request.form.getlist("publisher_ipi[]")
        mlcs = request.form.getlist("mlc_publisher_number[]")
        addresses = request.form.getlist("address[]")
        cities = request.form.getlist("city[]")
        states = request.form.getlist("state[]")
        zips = request.form.getlist("zip_code[]")
        emails = request.form.getlist("contact_email[]")
        phones = request.form.getlist("contact_phone[]")
        pids = request.form.getlist("pub_id[]")
        try:
            for i, name in enumerate(names):
                name = name.strip()
                if not name:
                    continue
                pid = pids[i] if i < len(pids) else ""
                if pid:
                    pc = PublisherConfig.query.get(int(pid))
                else:
                    pc = PublisherConfig.query.filter_by(publisher_name=name).first()
                    if not pc:
                        pc = PublisherConfig(publisher_name=name)
                        db.session.add(pc)
                pc.publisher_name = name
                pc.pro = pros[i].strip() if i < len(pros) else ""
                pc.publisher_ipi = ipis[i].strip() if i < len(ipis) else ""
                pc.mlc_publisher_number = mlcs[i].strip() if i < len(mlcs) else ""
                pc.address = addresses[i].strip() if i < len(addresses) else ""
                pc.city = cities[i].strip() if i < len(cities) else ""
                pc.state = states[i].strip() if i < len(states) else ""
                pc.zip_code = zips[i].strip() if i < len(zips) else ""
                pc.contact_email = emails[i].strip() if i < len(emails) else ""
                pc.contact_phone = phones[i].strip() if i < len(phones) else ""
            db.session.commit()
            flash("Publisher configuration saved.")
        except Exception as e:
            db.session.rollback()
            flash("Error saving: " + str(e))
        return redirect(url_for("reports.publisher_config"))

    # Pre-populate with Afinarte publishers if not yet configured
    existing_names = {c.publisher_name for c in PublisherConfig.query.all()}
    for ap in AFINARTE_PUBLISHERS:
        if ap not in existing_names:
            db.session.add(PublisherConfig(publisher_name=ap))
    if len(existing_names) < len(AFINARTE_PUBLISHERS):
        db.session.commit()

    configs = PublisherConfig.query.order_by(PublisherConfig.publisher_name).all()
    return render_template_string(PUBLISHER_CONFIG_HTML, configs=configs)


# ── Works Registration (consolidated PRO + Mechanical pipeline) ───────────────

@bp.route("/pro-registration")
def pro_registration_redirect():
    """Legacy redirect — keep old bookmarks working."""
    tab = request.args.get("tab", "unregistered")
    tab_map = {"registered": "submitted_to_pros", "unregistered": "unregistered"}
    return redirect(url_for("reports.works_registration", tab=tab_map.get(tab, tab)))


@bp.route("/works-registration")
def works_registration():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))

    tab = request.args.get("tab", "unregistered")
    q = (request.args.get("q") or "").strip()
    try:
        page = max(1, int(request.args.get("page") or 1))
    except (ValueError, TypeError):
        page = 1
    per_page = 50

    pro_registered_ids = db.session.query(ProRegistration.work_id).distinct()
    works_with_release = (db.session.query(TrackWork.work_id)
                          .join(Track, Track.id == TrackWork.track_id).distinct())

    # ── Unregistered ──────────────────────────────────────────────────────────
    unregistered_q = (
        Work.query
        .join(WorkWriter, WorkWriter.work_id == Work.id)
        .filter(WorkWriter.publisher.in_(AFINARTE_PUBLISHERS))
        .filter(Work.id.notin_(pro_registered_ids))
        .filter(Work.id.in_(works_with_release))
        .distinct()
        .order_by(Work.created_at.desc())
    )

    # ── Submitted to PROs ─────────────────────────────────────────────────────
    # Has ProRegistration row + at least one of MLC/MRI not yet submitted + not confirmed
    submitted_pros_q = (
        Work.query
        .join(ProRegistration, ProRegistration.work_id == Work.id)
        .filter(Work.id.in_(works_with_release))
        .filter(Work.registration_status != "confirmed")
        .filter(
            (Work.mlc_submitted_at == None) | (Work.mri_submitted_at == None)
        )
        .distinct()
        .order_by(Work.created_at.desc())
    )

    # ── Submitted to MLC ──────────────────────────────────────────────────────
    submitted_mlc_q = (
        Work.query
        .filter(Work.mlc_submitted_at != None)
        .filter(Work.registration_status != "confirmed")
        .order_by(Work.mlc_submitted_at.desc())
    )

    # ── Submitted to MRI ──────────────────────────────────────────────────────
    submitted_mri_q = (
        Work.query
        .filter(Work.mri_submitted_at != None)
        .filter(Work.registration_status != "confirmed")
        .order_by(Work.mri_submitted_at.desc())
    )

    # ── Confirmed ─────────────────────────────────────────────────────────────
    confirmed_q = (
        Work.query
        .filter(Work.registration_status == "confirmed")
        .order_by(Work.title)
    )

    if q:
        like_q = f"%{q.lower()}%"
        unregistered_q    = unregistered_q.filter(_func.lower(Work.title).like(like_q))
        submitted_pros_q  = submitted_pros_q.filter(_func.lower(Work.title).like(like_q))
        submitted_mlc_q   = submitted_mlc_q.filter(_func.lower(Work.title).like(like_q))
        submitted_mri_q   = submitted_mri_q.filter(_func.lower(Work.title).like(like_q))
        confirmed_q       = confirmed_q.filter(_func.lower(Work.title).like(like_q))

    # Counts for tab badges (always computed)
    unregistered_count   = unregistered_q.count()
    submitted_pros_count = submitted_pros_q.count()
    submitted_mlc_count  = submitted_mlc_q.count()
    submitted_mri_count  = submitted_mri_q.count()
    confirmed_count      = confirmed_q.count()

    # Paginate the active tab
    unregistered = submitted_pros = submitted_mlc = submitted_mri = confirmed = []
    pagination = None

    if tab == "submitted_to_pros":
        # No pagination — all works must be visible so one MLC/MRI export covers them all
        submitted_pros = submitted_pros_q.all()
        pagination = None
        reg_ids_map = [w.id for w in submitted_pros]
        all_regs = (ProRegistration.query
                    .filter(ProRegistration.work_id.in_(reg_ids_map))
                    .order_by(ProRegistration.registered_at.desc()).all())
        regs_by_work = {}
        for r in all_regs:
            regs_by_work.setdefault(r.work_id, []).append(r)
        for w in submitted_pros:
            w.registrations = regs_by_work.get(w.id, [])
        _attach_track_info_bulk(submitted_pros)

    elif tab == "submitted_to_mlc":
        pagination = submitted_mlc_q.paginate(page=page, per_page=per_page, error_out=False)
        submitted_mlc = pagination.items
        _attach_track_info_bulk(submitted_mlc)

    elif tab == "submitted_to_mri":
        pagination = submitted_mri_q.paginate(page=page, per_page=per_page, error_out=False)
        submitted_mri = pagination.items
        _attach_track_info_bulk(submitted_mri)

    elif tab == "confirmed":
        pagination = confirmed_q.paginate(page=page, per_page=per_page, error_out=False)
        confirmed = pagination.items
        _attach_track_info_bulk(confirmed)

    else:  # unregistered (default)
        tab = "unregistered"
        pagination = unregistered_q.paginate(page=page, per_page=per_page, error_out=False)
        unregistered = pagination.items
        _attach_track_info_bulk(unregistered)

    today = datetime.date.today().strftime("%Y-%m-%d")
    return render_template_string(WORKS_REGISTRATION_HTML,
        tab=tab, q=q, today=today, pagination=pagination,
        unregistered=unregistered,
        submitted_pros=submitted_pros,
        submitted_mlc=submitted_mlc,
        submitted_mri=submitted_mri,
        confirmed=confirmed,
        unregistered_count=unregistered_count,
        submitted_pros_count=submitted_pros_count,
        submitted_mlc_count=submitted_mlc_count,
        submitted_mri_count=submitted_mri_count,
        confirmed_count=confirmed_count,
    )


@bp.route("/works-registration/mark", methods=["POST"])
def works_registration_mark():
    """Create ProRegistration rows for selected works (PRO submission step)."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = request.form.getlist("work_ids[]")
    pro = request.form.get("pro", "").strip()
    pro_work_number = request.form.get("pro_work_number", "").strip()
    mlc_song_code = request.form.get("mlc_song_code", "").strip()
    registered_by = request.form.get("registered_by", "Omar").strip()
    registered_at_str = request.form.get("registered_at", "").strip()
    notes = request.form.get("notes", "").strip()
    try:
        registered_at = (datetime.datetime.strptime(registered_at_str, "%Y-%m-%d").date()
                         if registered_at_str else datetime.date.today())
    except ValueError:
        registered_at = datetime.date.today()
    if not pro or not work_ids:
        flash("Please select at least one work and a PRO.")
        return redirect(url_for("reports.works_registration"))
    try:
        for wid in work_ids:
            db.session.add(ProRegistration(
                work_id=int(wid), pro=pro, pro_work_number=pro_work_number,
                mlc_song_code=mlc_song_code, registered_at=registered_at,
                registered_by=registered_by, notes=notes,
            ))
        db.session.commit()
        flash(f"{len(work_ids)} work(s) marked as registered with {pro}.")
    except Exception as e:
        db.session.rollback()
        flash("Error: " + str(e))
    return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))


@bp.route("/works-registration/<int:reg_id>/delete", methods=["POST"])
def works_registration_delete(reg_id):
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    reg = ProRegistration.query.get_or_404(reg_id)
    db.session.delete(reg)
    db.session.commit()
    flash("Registration removed.")
    return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))


@bp.route("/works-registration/submit-mlc", methods=["POST"])
def works_registration_submit_mlc():
    """Mark selected works as submitted to MLC and download the MLC export."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = [int(x) for x in request.form.getlist("work_ids[]") if x]
    if not work_ids:
        flash("No works selected.")
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    now = datetime.datetime.utcnow()
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    for w in works:
        w.mlc_submitted_at = now
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash("Error saving: " + str(e))
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    # Generate and return the MLC export
    return _generate_mlc_export(work_ids)


@bp.route("/works-registration/submit-mri", methods=["POST"])
def works_registration_submit_mri():
    """Mark selected works as submitted to MRI and download the MRI export."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = [int(x) for x in request.form.getlist("work_ids[]") if x]
    if not work_ids:
        flash("No works selected.")
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    now = datetime.datetime.utcnow()
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    for w in works:
        w.mri_submitted_at = now
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash("Error saving: " + str(e))
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    return _generate_mri_export(work_ids)


@bp.route("/works-registration/submit-both", methods=["POST"])
def works_registration_submit_both():
    """Mark selected works as submitted to both MLC and MRI, return a zip with both exports."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = [int(x) for x in request.form.getlist("work_ids[]") if x]
    if not work_ids:
        flash("No works selected.")
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    now = datetime.datetime.utcnow()
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    for w in works:
        w.mlc_submitted_at = now
        w.mri_submitted_at = now
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash("Error saving: " + str(e))
        return redirect(url_for("reports.works_registration", tab="submitted_to_pros"))
    import zipfile
    mlc_bytes = _generate_mlc_bytes(work_ids)
    mri_bytes = _generate_mri_bytes(work_ids)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("MLC_submission.xlsx", mlc_bytes.getvalue())
        zf.writestr("MRI_submission.xls", mri_bytes.getvalue())
    zip_buf.seek(0)
    return send_file(zip_buf, mimetype="application/zip",
                     download_name="mechanical_submission.zip", as_attachment=True)


@bp.route("/works-registration/resubmit-mlc", methods=["POST"])
def works_registration_resubmit_mlc():
    """Clear mlc_submitted_at so the work returns to Submitted to PROs for re-export."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = [int(x) for x in request.form.getlist("work_ids[]") if x]
    if not work_ids:
        flash("No works selected.")
        return redirect(url_for("reports.works_registration", tab="submitted_to_mlc"))
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    for w in works:
        w.mlc_submitted_at = None
    try:
        db.session.commit()
        flash(f"{len(works)} work(s) moved back to Submitted to PROs for MLC resubmission.")
    except Exception as e:
        db.session.rollback()
        flash("Error: " + str(e))
    return redirect(url_for("reports.works_registration", tab="submitted_to_mlc"))


@bp.route("/works-registration/resubmit-mri", methods=["POST"])
def works_registration_resubmit_mri():
    """Clear mri_submitted_at so the work returns to Submitted to PROs for re-export."""
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_ids = [int(x) for x in request.form.getlist("work_ids[]") if x]
    if not work_ids:
        flash("No works selected.")
        return redirect(url_for("reports.works_registration", tab="submitted_to_mri"))
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    for w in works:
        w.mri_submitted_at = None
    try:
        db.session.commit()
        flash(f"{len(works)} work(s) moved back to Submitted to PROs for MRI resubmission.")
    except Exception as e:
        db.session.rollback()
        flash("Error: " + str(e))
    return redirect(url_for("reports.works_registration", tab="submitted_to_mri"))


# ── Export helpers (used by submit-mlc / submit-mri / submit-both) ────────────

def _track_info_for_export(work_id):
    """Return (rec_title, rec_artist, rec_isrc, rec_label, upc) for first linked track."""
    tracks = (Track.query
              .join(TrackWork, TrackWork.track_id == Track.id)
              .filter(TrackWork.work_id == work_id)
              .all())
    if not tracks:
        return "", "", "", "", ""
    t = tracks[0]
    try:
        rec_artist = ", ".join(a for a in _json.loads(t.artists or "[]") if a)
    except Exception:
        rec_artist = t.artists or ""
    upc = t.release.upc if t.release else ""
    return t.primary_title or "", rec_artist, t.isrc or "", t.track_label or "", upc or ""


def _generate_mlc_bytes(work_ids):
    """Build MLC export workbook for the given work IDs. Returns BytesIO."""
    import openpyxl
    from openpyxl import load_workbook
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    template_path = os.path.join(_TEMPLATE_DIR, "MLCBulkWork_V1.2-2.xlsx")
    wb = load_workbook(template_path)
    ws = wb["Format"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    row_idx = 2
    for work in works:
        wws = WorkWriter.query.filter_by(work_id=work.id).all()
        rec_title, rec_artist, rec_isrc, rec_label, _ = _track_info_for_export(work.id)
        rec_isrc = (rec_isrc or "").replace("-", "") or None
        first_writer = True
        for ww in wws:
            wr = ww.writer
            ws.cell(row=row_idx, column=1).value  = work.title if first_writer else None
            ws.cell(row=row_idx, column=2).value  = None
            ws.cell(row=row_idx, column=3).value  = f"LM{work.id:06d}"
            ws.cell(row=row_idx, column=4).value  = work.iswc or None
            ws.cell(row=row_idx, column=5).value  = work.aka_title or None
            ws.cell(row=row_idx, column=6).value  = work.aka_title_type_code or None
            ws.cell(row=row_idx, column=7).value  = wr.last_names or None
            ws.cell(row=row_idx, column=8).value  = wr.first_name or None
            ws.cell(row=row_idx, column=9).value  = wr.ipi or None
            ws.cell(row=row_idx, column=10).value = ww.writer_role_code or "CA"
            ws.cell(row=row_idx, column=11).value = None
            ws.cell(row=row_idx, column=12).value = ww.publisher or None
            ws.cell(row=row_idx, column=13).value = ww.publisher_ipi or None
            ws.cell(row=row_idx, column=14).value = None
            ws.cell(row=row_idx, column=15).value = ww.administrator_name or None
            ws.cell(row=row_idx, column=16).value = ww.administrator_ipi or None
            ws.cell(row=row_idx, column=17).value = ww.writer_percentage or None
            ws.cell(row=row_idx, column=18).value = rec_title or None
            ws.cell(row=row_idx, column=19).value = rec_artist or None
            ws.cell(row=row_idx, column=20).value = rec_isrc or None
            ws.cell(row=row_idx, column=21).value = rec_label or None
            row_idx += 1
            first_writer = False
        if not wws:
            ws.cell(row=row_idx, column=1).value  = work.title
            ws.cell(row=row_idx, column=3).value  = f"LM{work.id:06d}"
            ws.cell(row=row_idx, column=4).value  = work.iswc or None
            ws.cell(row=row_idx, column=18).value = rec_title or None
            ws.cell(row=row_idx, column=19).value = rec_artist or None
            ws.cell(row=row_idx, column=20).value = rec_isrc or None
            row_idx += 1
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate_mri_bytes(work_ids):
    """Build MRI export workbook for the given work IDs. Returns BytesIO."""
    from blueprints.export_helpers import open_xls_template
    works = Work.query.filter(Work.id.in_(work_ids)).all()
    template_path = os.path.join(_TEMPLATE_DIR, "MusicReportspublishing_catalog_template-3.xls")
    wb_xls = open_xls_template(template_path)
    ws = wb_xls.get_sheet(0)
    row_idx = 1
    for work in works:
        wws = WorkWriter.query.filter_by(work_id=work.id).all()
        rec_title, rec_artist, rec_isrc, rec_label, upc = _track_info_for_export(work.id)
        rec_isrc = (rec_isrc or "").replace("-", "")
        first_writer = True
        for ww in wws:
            wr = ww.writer
            ws.write(row_idx, 0,  work.title if first_writer else "")
            ws.write(row_idx, 1,  work.aka_title or "")
            ws.write(row_idx, 2,  work.mri_song_id or "")
            ws.write(row_idx, 3,  f"LM{work.id:06d}")
            ws.write(row_idx, 4,  work.iswc or "")
            ws.write(row_idx, 5,  wr.last_names or "")
            ws.write(row_idx, 6,  wr.first_name or "")
            ws.write(row_idx, 7,  wr.middle_name or "")
            ws.write(row_idx, 8,  wr.pro or "")
            ws.write(row_idx, 9,  wr.ipi or "")
            ws.write(row_idx, 10, "Y")
            ws.write(row_idx, 11, ww.writer_percentage or 0)
            ws.write(row_idx, 12, ww.writer_role_code or "CA")
            ws.write(row_idx, 13, ww.publisher or "")
            ws.write(row_idx, 14, "")
            ws.write(row_idx, 15, ww.publisher_ipi or "")
            ws.write(row_idx, 16, "Y")
            ws.write(row_idx, 17, ww.administrator_name or "")
            ws.write(row_idx, 18, ww.writer_percentage or 0)
            ws.write(row_idx, 19, ww.territory_controlled or "World")
            ws.write(row_idx, 20, "")
            ws.write(row_idx, 21, "")
            ws.write(row_idx, 22, "")
            ws.write(row_idx, 23, rec_artist if first_writer else "")
            ws.write(row_idx, 24, rec_label  if first_writer else "")
            ws.write(row_idx, 25, rec_isrc   if first_writer else "")
            ws.write(row_idx, 26, upc        if first_writer else "")
            row_idx += 1
            first_writer = False
        if not wws:
            ws.write(row_idx, 0, work.title)
            ws.write(row_idx, 2, work.mri_song_id or "")
            ws.write(row_idx, 3, f"LM{work.id:06d}")
            ws.write(row_idx, 4, work.iswc or "")
            row_idx += 1
    buf = io.BytesIO()
    wb_xls.save(buf)
    buf.seek(0)
    return buf


def _generate_mlc_export(work_ids):
    buf = _generate_mlc_bytes(work_ids)
    filename = f"MLC_Submission_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name=filename, as_attachment=True)


def _generate_mri_export(work_ids):
    buf = _generate_mri_bytes(work_ids)
    filename = f"MRI_Submission_{datetime.date.today().strftime('%Y%m%d')}.xls"
    return send_file(buf, mimetype="application/vnd.ms-excel",
                     download_name=filename, as_attachment=True)


# ── Reports Index ─────────────────────────────────────────────────────────────

@bp.route("/reports")
def reports_index():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    work_count = (Work.query
                  .join(WorkWriter, WorkWriter.work_id == Work.id)
                  .filter(WorkWriter.publisher.in_(AFINARTE_PUBLISHERS))
                  .distinct().count())
    release_count = Release.query.count()
    return render_template_string(REPORTS_INDEX_HTML, work_count=work_count, release_count=release_count)


# ── MLC Export ───────────────────────────────────────────────────────────────

@bp.route("/reports/export/mlc")
def export_mlc():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        import openpyxl
        from openpyxl import load_workbook

        wb = load_workbook("template/MLCBulkWork_V1.2-2.xlsx")
        ws = wb["Format"]

        # Clear example rows (keep header row 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        works = (Work.query
                 .join(WorkWriter, WorkWriter.work_id == Work.id)
                 .filter(WorkWriter.publisher.in_(AFINARTE_PUBLISHERS))
                 .distinct()
                 .order_by(Work.title)
                 .all())

        row_idx = 2
        for work in works:
            writers = WorkWriter.query.filter_by(work_id=work.id).all()
            first_writer = True
            for ww in writers:
                w = ww.writer
                pub_config = PublisherConfig.query.filter(
                    _func.lower(PublisherConfig.publisher_name) == (ww.publisher or "").lower()
                ).first()
                mlc_pub_num = pub_config.mlc_publisher_number if pub_config else ""

                tracks = (Track.query
                          .join(TrackWork, TrackWork.track_id == Track.id)
                          .filter(TrackWork.work_id == work.id)
                          .all())
                rec_title = tracks[0].primary_title if tracks else ""
                rec_artist = ""
                if tracks:
                    try:
                        al = _json.loads(tracks[0].artists or "[]")
                        rec_artist = ", ".join(al)
                    except Exception:
                        pass
                rec_isrc = (tracks[0].isrc or "").replace("-", "") if tracks else ""
                rec_label = tracks[0].track_label if tracks else ""

                ws.cell(row=row_idx, column=1).value = work.title if first_writer else None
                ws.cell(row=row_idx, column=2).value = None  # MLC Song Code
                ws.cell(row=row_idx, column=3).value = f"LM{work.id:06d}"
                ws.cell(row=row_idx, column=4).value = work.iswc or None
                ws.cell(row=row_idx, column=5).value = work.aka_title or None
                ws.cell(row=row_idx, column=6).value = work.aka_title_type_code or None
                ws.cell(row=row_idx, column=7).value = w.last_names
                ws.cell(row=row_idx, column=8).value = w.first_name
                ws.cell(row=row_idx, column=9).value = w.ipi or None
                ws.cell(row=row_idx, column=10).value = ww.writer_role_code or "CA"
                ws.cell(row=row_idx, column=11).value = mlc_pub_num or None
                ws.cell(row=row_idx, column=12).value = ww.publisher or None
                ws.cell(row=row_idx, column=13).value = ww.publisher_ipi or None
                ws.cell(row=row_idx, column=14).value = None
                ws.cell(row=row_idx, column=15).value = ww.administrator_name or None
                ws.cell(row=row_idx, column=16).value = ww.administrator_ipi or None
                ws.cell(row=row_idx, column=17).value = ww.writer_percentage or None
                ws.cell(row=row_idx, column=18).value = rec_title or None
                ws.cell(row=row_idx, column=19).value = rec_artist or None
                ws.cell(row=row_idx, column=20).value = rec_isrc or None
                ws.cell(row=row_idx, column=21).value = rec_label or None

                row_idx += 1
                first_writer = False

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"MLC_BulkWork_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(output, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True)
    except Exception as e:
        current_app.logger.error("MLC export error: %s", e)
        import traceback; current_app.logger.error(traceback.format_exc())
        flash("Error generating MLC export: " + str(e))
        return redirect(url_for("reports.reports_index"))


# ── Music Reports Export ──────────────────────────────────────────────────────

@bp.route("/reports/export/music-reports")
def export_music_reports():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        from blueprints.export_helpers import open_xls_template

        template_path = os.path.join(_TEMPLATE_DIR, "MusicReportspublishing_catalog_template-3.xls")
        wb = open_xls_template(template_path)
        ws = wb.get_sheet(0)  # row 0 = header, preserved from template

        works = (Work.query
                 .join(WorkWriter, WorkWriter.work_id == Work.id)
                 .filter(WorkWriter.publisher.in_(AFINARTE_PUBLISHERS))
                 .distinct()
                 .order_by(Work.title)
                 .all())

        row_idx = 1
        for work in works:
            writers = WorkWriter.query.filter_by(work_id=work.id).all()
            tracks = (Track.query
                      .join(TrackWork, TrackWork.track_id == Track.id)
                      .filter(TrackWork.work_id == work.id)
                      .all())
            rec_artist = rec_isrc = rec_label = upc = ""
            if tracks:
                try:
                    al = _json.loads(tracks[0].artists or "[]")
                    rec_artist = ", ".join(al)
                except Exception:
                    pass
                rec_isrc = (tracks[0].isrc or "").replace("-", "")
                rec_label = tracks[0].track_label or ""
                if tracks[0].release:
                    upc = tracks[0].release.upc or ""

            first_writer = True
            for ww in writers:
                w = ww.writer
                pub_config = PublisherConfig.query.filter(
                    _func.lower(PublisherConfig.publisher_name) == (ww.publisher or "").lower()
                ).first()
                pub_address = pub_contact = pub_pro = ""
                if pub_config:
                    parts = [pub_config.address, pub_config.city]
                    if pub_config.state:
                        parts.append(pub_config.state)
                    if pub_config.zip_code:
                        parts.append(pub_config.zip_code)
                    pub_address = ", ".join(p for p in parts if p)
                    pub_contact = pub_config.contact_email or pub_config.contact_phone or ""
                    pub_pro = pub_config.pro or ""

                controlled = "Y" if _is_controlled(ww.publisher) else "N"

                ws.write(row_idx, 0, work.title if first_writer else "")
                ws.write(row_idx, 1, work.aka_title or "")
                ws.write(row_idx, 2, work.mri_song_id or "")
                ws.write(row_idx, 3, f"LM{work.id:06d}")
                ws.write(row_idx, 4, work.iswc or "")
                ws.write(row_idx, 5, w.last_names or "")
                ws.write(row_idx, 6, w.first_name or "")
                ws.write(row_idx, 7, w.middle_name or "")
                ws.write(row_idx, 8, w.pro or "")
                ws.write(row_idx, 9, w.ipi or "")
                ws.write(row_idx, 10, controlled)
                ws.write(row_idx, 11, ww.writer_percentage or 0)
                ws.write(row_idx, 12, ww.writer_role_code or "CA")
                ws.write(row_idx, 13, ww.publisher or "")
                ws.write(row_idx, 14, pub_pro)
                ws.write(row_idx, 15, ww.publisher_ipi or "")
                ws.write(row_idx, 16, controlled)
                ws.write(row_idx, 17, ww.administrator_name or "")
                ws.write(row_idx, 18, ww.writer_percentage or 0)
                ws.write(row_idx, 19, ww.territory_controlled or "World")
                ws.write(row_idx, 20, "")
                ws.write(row_idx, 21, pub_address)
                ws.write(row_idx, 22, pub_contact)
                ws.write(row_idx, 23, rec_artist if first_writer else "")
                ws.write(row_idx, 24, rec_label if first_writer else "")
                ws.write(row_idx, 25, rec_isrc if first_writer else "")
                ws.write(row_idx, 26, upc if first_writer else "")

                row_idx += 1
                first_writer = False

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"MusicReports_{datetime.date.today().strftime('%Y%m%d')}.xls"
        return send_file(output, download_name=filename,
                         mimetype="application/vnd.ms-excel",
                         as_attachment=True)
    except Exception as e:
        current_app.logger.error("Music Reports export error: %s", e)
        import traceback; current_app.logger.error(traceback.format_exc())
        flash("Error generating Music Reports export: " + str(e))
        return redirect(url_for("reports.reports_index"))


# ── SoundExchange Export ──────────────────────────────────────────────────────

@bp.route("/reports/export/soundexchange")
def export_soundexchange():
    if auth_required():
        return redirect(url_for("publishing.login"))
    if role_required(FULL_ACCESS_ROLES):
        flash("Access restricted.", "error")
        return redirect(url_for("publishing.works_list"))
    try:
        from openpyxl import load_workbook
        from blueprints.export_helpers import stitch_xlsx_assets

        sx_template_path = os.path.join(_TEMPLATE_DIR, "Sound Exchange ISRC Ingest Form.xlsx")
        wb = load_workbook(sx_template_path)
        ws = wb["Form"]

        # Clear data below header (row 10 is header, data starts row 11)
        data_start = 11
        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        tracks = (Track.query
                  .join(Release, Release.id == Track.release_id)
                  .order_by(Release.title, Track.track_number)
                  .all())

        row_idx = data_start
        for t in tracks:
            artist = ""
            try:
                al = _json.loads(t.artists or "[]")
                artist = ", ".join(al)
            except Exception:
                pass
            if not artist and t.release:
                try:
                    ral = _json.loads(t.release.artists or "[]")
                    artist = ", ".join(ral)
                except Exception:
                    pass

            ws.cell(row=row_idx, column=1).value = artist
            ws.cell(row=row_idx, column=2).value = t.primary_title
            ws.cell(row=row_idx, column=3).value = t.isrc or ""
            ws.cell(row=row_idx, column=4).value = "Copyright Owner"
            ws.cell(row=row_idx, column=5).value = 100
            ws.cell(row=row_idx, column=6).value = (t.release.release_date.strftime("%m/%d/%Y")
                                                     if t.release and t.release.release_date else "")
            ws.cell(row=row_idx, column=7).value = ""
            ws.cell(row=row_idx, column=8).value = ""
            ws.cell(row=row_idx, column=9).value = ""
            ws.cell(row=row_idx, column=10).value = t.duration or ""
            ws.cell(row=row_idx, column=11).value = t.genre or ""
            ws.cell(row=row_idx, column=12).value = (t.recording_date.strftime("%m/%d/%Y")
                                                      if t.recording_date else "")
            ws.cell(row=row_idx, column=13).value = t.country_of_recording or "US"
            ws.cell(row=row_idx, column=14).value = ""
            ws.cell(row=row_idx, column=15).value = "US"
            row_idx += 1

        opx_buf = io.BytesIO()
        wb.save(opx_buf)
        fixed_bytes = stitch_xlsx_assets(sx_template_path, opx_buf.getvalue())
        filename = f"SoundExchange_ISRC_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return send_file(io.BytesIO(fixed_bytes), download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True)
    except Exception as e:
        current_app.logger.error("SoundExchange export error: %s", e)
        import traceback; current_app.logger.error(traceback.format_exc())
        flash("Error generating SoundExchange export: " + str(e))
        return redirect(url_for("reports.reports_index"))
